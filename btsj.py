import glob
import re
from openpyxl import load_workbook

data = []

for xlsx in glob.glob('btsjcorpus_ver2020/**/*-*-*-*.xlsx', recursive=True):
    wb = load_workbook(xlsx)
    ws = wb.active

    for row in ws.iter_rows(min_row=4, values_only=True):
        cell_range = row[3:8]
        if None in cell_range:
            break
        data.append(tuple(map(str, cell_range)))

    wb.close()

symbol = re.compile(r'‘.*?’|\[.*?\]|《.*?》|=|【【|】】|\(.*?\)|<.*?>|(<.*?>)')
pairs = []

for i in range(len(data) - 1):
    if (
        data[i][2] == '/' or data[i+1][2] == '/'
        or data[i][3] == data[i+1][3]
        or '{<}' in data[i][4] or '{>}' in data[i][4]
        or '{<}' in data[i+1][4] or '{>}' in data[i+1][4]
        or '#' in data[i][4] or '#' in data[i+1][4]
    ):
        continue

    src = symbol.sub('', data[i][4])
    tgt = symbol.sub('', data[i+1][4])

    if src == '。' or tgt == '。':
        continue

    pairs.append((src, tgt))

with open('btsj.tsv', 'w') as f:
    for pair in pairs:
        f.write('\t'.join(pair) + '\n')
